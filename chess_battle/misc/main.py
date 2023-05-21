import random
from openpyxl import Workbook
import sqlite3


conn = sqlite3.connect("battle_data.sqlite3")

cursor = conn.cursor()


"""
# chess player distribution
1. In the first run, players was setted down randomly
2. From 2nd round, players distribute by their accumulated scores

table view
-----------------------------------------------------------------
SCORE TABLE
===========

NO.     NAME    SCORE
####    ####    ####
1       MAX     0
2       JESS    0
3       HACK    0


PLAYER BATTLES GROUPS TABLE
===========================

NO.     BALTE_A     BALTE_B
====    ====        ====
1       2           30
2       4           28
"""

# Each player should conclude a number and each round their scores
players = [{x: [0]} for x in range(30)]
print(players)

number_players = len(players)

wb_name = "battles.xlsx"
wb = Workbook()

headers_score = """
序号    姓名    分数
====    ====    ====
"""
print(headers_score)

ws = wb.active
ws.cell(1, 1, "No")
ws.cell(1, 2, "Name")
ws.cell(1, 3, "Score")
for item in range(number_players):
    player_name = list(players[item].keys())[0]
    player_score = players[item][player_name][0]
    print(
        "{:0>2d}      {:0>2d}      {:0>2d}".format(
            item + 1, player_name, player_score)
    )
    cursor.execute(
        "INSERT INTO player(ID, Name, Age) VALUES(?,?,?)",
        (item + 1, player_name + 1, player_score),
    )
    ws.cell(item + 2, 1, item + 1)
    ws.cell(item + 2, 2, player_name + 1)
    ws.cell(item + 2, 3, player_score)

cursor.close()
conn.commit()
conn.close()
headers_battle = """
序号    先出    后出
====    ====    ====
"""
print(headers_battle)
ws.cell(1, 5, "No")
ws.cell(1, 6, "First")
ws.cell(1, 7, "Last")
for x in range(number_players // 2):
    choice1 = random.choice(players)
    players.remove(choice1)
    choice2 = random.choice(players)
    players.remove(choice2)
    print(
        "{:0>2d}      {:0>2d}      {:0>2d}".format(
            x + 1, list(choice1.keys())[0], list(choice2.keys())[0]
        )
    )

    ws.cell(x + 2, 5, x + 1)
    ws.cell(x + 2, 6, list(choice1.keys())[0])
    ws.cell(x + 2, 7, list(choice2.keys())[0])

wb.save(wb_name)
